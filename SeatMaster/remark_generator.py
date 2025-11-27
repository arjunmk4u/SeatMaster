import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font
import os
import glob

def build_student_mapping(students_dir="data/students"):
    """
    Scans Excel files in students_dir to create a {Class_No: Batch_Name} map.
    Looks for 'Batch Name' and 'Class No' inside the files.
    """
    mapping = {}
    
    if not os.path.exists(students_dir):
        print(f"Warning: Directory '{students_dir}' not found.")
        return mapping

    print(f"Scanning student files in {students_dir}...")
    files = glob.glob(os.path.join(students_dir, "*.xlsx"))
    
    for file_path in files:
        try:
            # 1. Scan first 15 rows to find the header row dynamically
            df_scan = pd.read_excel(file_path, header=None, nrows=15)
            
            header_row_idx = None
            for idx, row in df_scan.iterrows():
                row_str = [str(val).strip() for val in row.values]
                if "Class No" in row_str and "Batch Name" in row_str:
                    header_row_idx = idx
                    break
            
            # 2. Read file with correct header
            if header_row_idx is not None:
                df = pd.read_excel(file_path, header=header_row_idx)
                # Normalize columns
                df.columns = [str(c).strip() for c in df.columns]
                
                if "Class No" in df.columns and "Batch Name" in df.columns:
                    df = df.dropna(subset=["Class No"])
                    for _, row in df.iterrows():
                        c_no = str(row["Class No"]).strip()
                        b_name = str(row["Batch Name"]).strip()
                        mapping[c_no] = b_name
                        
        except Exception as e:
            print(f"Error processing {os.path.basename(file_path)}: {e}")
            
    print(f"Mapped {len(mapping)} students to their classes.")
    return mapping

def generate_remark_sheets(seating_df, exam_title, exam_date, template_path, output_path, students_dir="data/students"):
    """
    Generates remark sheets. 
    Updates: Bold Seat No, White Footer Headers, and internal Class Name mapping.
    """
    
    # 1. Build Mapping
    class_map = build_student_mapping(students_dir)
    
    # 2. Setup
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    try:
        wb = openpyxl.load_workbook(template_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"Template not found at: {template_path}")
    
    master_sheet = wb.active
    master_sheet.title = "Master_Template"

    grouped = seating_df.groupby("Room")

    for room_name, room_data in grouped:
        ws = wb.copy_worksheet(master_sheet)
        ws.title = str(room_name)[:30]
        
        # --- Header ---
        ws['B3'] = exam_date
        ws['F3'] = str(room_name)      
        if ws['A2'].value and "Examination" in str(ws['A2'].value):
             ws['A2'] = exam_title
        elif ws['A1'].value:
             ws['A1'] = exam_title

        # --- Row Expansion Logic ---
        START_ROW = 7
        
        # Find Footer Header Row (starts with "Class")
        footer_header_row = 25
        for r in range(START_ROW, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val and str(val).strip() == "Class":
                footer_header_row = r
                break
        
        room_data['Bench'] = pd.to_numeric(room_data['Bench'], errors='coerce').fillna(0).astype(int)
        min_bench = room_data['Bench'].min()
        max_bench = room_data['Bench'].max()
        
        needed_rows = (max_bench - min_bench) + 1
        available_rows = footer_header_row - START_ROW
        
        if needed_rows > available_rows:
            rows_to_add = needed_rows - available_rows + 2
            ws.insert_rows(footer_header_row, amount=rows_to_add)
            # Update the location of the footer header because it moved down
            footer_header_row += rows_to_add 

        # --- Fill Seating Data ---
        for bench_num in range(min_bench, max_bench + 1):
            current_row = START_ROW + (bench_num - min_bench)
            
            # Seat No (Col A) - BOLD
            c_seat = ws.cell(row=current_row, column=1, value=bench_num)
            c_seat.alignment = Alignment(horizontal='center', vertical='center')
            c_seat.font = Font(bold=True)  # <--- NEW: BOLD FONT

            bench_students = room_data[room_data['Bench'] == bench_num]
            
            for _, student in bench_students.iterrows():
                seat_pos = student['Seat']
                class_no = str(student['Class No']).strip()
                
                if class_no == "-" or not class_no or class_no == "nan":
                    continue

                target_col = None
                if seat_pos == "Left": target_col = 2
                elif seat_pos == "Center": target_col = 4
                elif seat_pos == "Right": target_col = 6
                
                if target_col:
                    c_class = ws.cell(row=current_row, column=target_col, value=class_no)
                    c_class.alignment = Alignment(horizontal='center', vertical='center')

        # --- Fill Summary Section ---
        
        # 1. Style the Footer Header (Class, Start, End, Count) -> WHITE TEXT
        # Columns 1 to 4 in the footer_header_row
        for col_idx in range(1, 5):
            header_cell = ws.cell(row=footer_header_row, column=col_idx)
            # Maintain existing bold/size if any, just enforce White Color
            # If you want to force Bold + White: Font(bold=True, color="FFFFFF")
            header_cell.font = Font(color="FFFFFF", bold=True) # <--- NEW: WHITE TEXT

        # 2. Prepare Data
        room_data['Real_Class_Name'] = room_data['Class No'].astype(str).str.strip().map(class_map).fillna("Unknown Class")
        valid_students = room_data[room_data['Class No'] != "-"]
        
        # 3. Total Students
        total_count = len(valid_students)
        # Scan slightly below the header to find "Total Students" label
        for r in range(footer_header_row, footer_header_row + 15):
            for c in range(1, 10):
                val = ws.cell(row=r, column=c).value
                if val and "Total Students" in str(val):
                    ws.cell(row=r, column=c+1, value=total_count)
                    break

        # 4. Summary Table Rows
        summary_stats = valid_students.groupby('Real_Class_Name')['Class No'].agg(['count', 'min', 'max']).reset_index()
        
        stat_row_idx = footer_header_row + 1 # Start writing below the header
        
        for _, row in summary_stats.iterrows():
            # A: Class Name
            c1 = ws.cell(row=stat_row_idx, column=1, value=row['Real_Class_Name'])
            c1.alignment = Alignment(horizontal='left', wrap_text=True, vertical='center')
            
            # B: Start
            c2 = ws.cell(row=stat_row_idx, column=2, value=row['min'])
            c2.alignment = Alignment(horizontal='center', vertical='center')
            
            # C: End
            c3 = ws.cell(row=stat_row_idx, column=3, value=row['max'])
            c3.alignment = Alignment(horizontal='center', vertical='center')
            
            # D: Count
            c4 = ws.cell(row=stat_row_idx, column=4, value=row['count'])
            c4.alignment = Alignment(horizontal='center', vertical='center')
            
            stat_row_idx += 1

    if "Master_Template" in wb.sheetnames:
        del wb["Master_Template"]
        
    wb.save(output_path)
    return output_path